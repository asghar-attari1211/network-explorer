import argparse
import pandas as pd
import re
from datetime import datetime
import gc
import os
import resource
from collections import defaultdict, Counter
from folium.plugins import Search
import folium
import openpyxl

parser = argparse.ArgumentParser(description='Microwave route tracing with sample and verbose modes')
parser.add_argument('--sample-sites', type=int, default=0, help='Limit to the first N SL sites for validation')
parser.add_argument('--sample-rows', type=int, default=0, help='Limit VLAN DB rows to the first N rows for validation')
parser.add_argument('--filter-sites', type=str, default='', help='Comma-separated prefixes to filter sites (e.g., KHD,IUCT)')
parser.add_argument('--dry-run', action='store_true', help='Skip writing Excel output file')
parser.add_argument('--verbose', action='store_true', help='Print detailed progress logs')
parser.add_argument('--debug-site', type=str, default='', help='Show detailed trace output for a specific site')
parser.add_argument('--no-map', action='store_true', help='Skip generating HTML map')
args = parser.parse_args()

DEBUG_SITE = args.debug_site.strip().upper() if args.debug_site else ''
FILE_PATH = "/home/asghar_attari1211/RnD/VLAN List -- AI Seekho -- Sample.xlsx"
SITE_PATTERN = re.compile(r'([A-Z]{2,10}\d{1,6})')


def extract_sites(text):
    if pd.isna(text):
        return []
    matches = SITE_PATTERN.findall(str(text).upper())
    return [m for m in matches if len(m) >= 6]


# Consolidate VLAN Reports
vlan_reports_folder = "/home/asghar_attari1211/RnD/VLAN List Reports"
consolidated_dfs = []
skipped_files = []

def extract_date_from_filename(fname):
    # look for pattern DD-MMM-YYYY anywhere in filename (case-insensitive)
    m = re.search(r"(\d{1,2}-[A-Za-z]{3}-\d{4})", fname)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1), '%d-%b-%Y').date()
    except Exception:
        return None

for file_name in os.listdir(vlan_reports_folder):
    if not file_name.endswith('.xlsx'):
        continue
    file_path = os.path.join(vlan_reports_folder, file_name)
    file_date = extract_date_from_filename(file_name)
    if not file_date:
        print(f"⚠️ VLAN file '{file_name}' does not contain a DD-MMM-YYYY date in its name. Please rename and re-run.")
        skipped_files.append(file_name)
        continue
    try:
        xl = pd.ExcelFile(file_path)
        sheet_names = xl.sheet_names
        vlan_sheet = None
        for sn in sheet_names:
            if 'VLAN' in sn.upper() or 'REPORT' in sn.upper():
                vlan_sheet = sn
                break
        if not vlan_sheet:
            vlan_sheet = sheet_names[0]
        # Try header=0 then header=1
        df = pd.read_excel(file_path, sheet_name=vlan_sheet, header=0)
        if 'NE NAME' not in df.columns or 'VLAN ID' not in df.columns:
            df = pd.read_excel(file_path, sheet_name=vlan_sheet, header=1)
        if not df.empty:
            # enforce DATE from filename
            df['DATE'] = pd.to_datetime(file_date).strftime('%d-%b-%Y')
            consolidated_dfs.append(df)
    except Exception as e:
        print(f"Error reading {file_name}: {e}")

if consolidated_dfs:
    vlan_report = pd.concat(consolidated_dfs, ignore_index=True)
    # Create combo for dedupe
    if 'NE NAME' in vlan_report.columns and 'VLAN Service Name' in vlan_report.columns and 'VLAN ID' in vlan_report.columns:
        vlan_report['combo'] = vlan_report.apply(lambda row: '-'.join(sorted([str(row['NE NAME']), str(row['VLAN Service Name'])])) + '-' + str(row['VLAN ID']), axis=1)
        vlan_report['DATE'] = pd.to_datetime(vlan_report['DATE'], errors='coerce')
        vlan_report = vlan_report.sort_values('DATE', ascending=False).drop_duplicates(subset='combo', keep='first')
        vlan_report = vlan_report.drop(columns=['combo'])
    # ensure DATE is formatted as DD-MMM-YYYY text
    try:
        vlan_report['DATE'] = pd.to_datetime(vlan_report['DATE'], errors='coerce').dt.strftime('%d-%b-%Y').fillna('')
    except Exception:
        vlan_report['DATE'] = vlan_report['DATE'].astype(str).fillna('')
else:
    # Fallback to main file, find correct sheet
    # If there are no consolidated reports in the VLAN Reports folder, inform the user
    try:
        report_files = [f for f in os.listdir(vlan_reports_folder) if f.endswith('.xlsx')]
    except Exception:
        report_files = []
    if not report_files:
        print(f"⚠️ No VLAN report files found in '{vlan_reports_folder}'. Please upload the latest VLAN Report (.xlsx) and re-run.")

    xl = pd.ExcelFile(FILE_PATH)
    sheet_names = xl.sheet_names
    vlan_sheet = None
    hdr = 1
    expected_cols = {'NE NAME', 'VLAN ID', 'VLAN SERVICE NAME', 'DATE'}
    # Try to locate a sheet that contains the expected columns (try header=0 and header=1)
    for sn in sheet_names:
        for header_try in (0, 1):
            try:
                tmp = pd.read_excel(FILE_PATH, sheet_name=sn, header=header_try, nrows=0)
                cols = {str(c).strip().upper() for c in tmp.columns}
                if expected_cols.issubset(cols):
                    vlan_sheet = sn
                    hdr = header_try
                    break
            except Exception:
                continue
        if vlan_sheet:
            break

    # If we didn't find a sheet by columns, fallback to name-based selection (existing behavior)
    if not vlan_sheet:
        for sn in sheet_names:
            if 'VLAN' in sn.upper() or 'REPORT' in sn.upper():
                vlan_sheet = sn
                hdr = 1
                break
    if not vlan_sheet:
        vlan_sheet = sheet_names[0]
        hdr = 1

    # Try reading with the expected usecols; if that fails, read without usecols and proceed
    try:
        vlan_report = pd.read_excel(FILE_PATH, sheet_name=vlan_sheet, header=hdr, usecols=['NE NAME', 'VLAN ID', 'VLAN Service Name', 'DATE'])
    except ValueError:
        print(f"⚠️ Expected VLAN report columns not found in sheet '{vlan_sheet}'. Reading sheet without usecols and attempting to adapt.")
        vlan_report = pd.read_excel(FILE_PATH, sheet_name=vlan_sheet, header=hdr)

print('🚀 [1/10] Loading workbook and building intelligence maps...')

vlan_report.columns = [str(c).strip() for c in vlan_report.columns]
# Use current date for output filenames (except consolidated VLAN report)
output_date = datetime.now().strftime('%d-%b-%Y')

# Defer saving consolidated VLAN report until export phase. If we have consolidated DFS,
# name the consolidated file using the latest DATE present in `vlan_report` (DD-MMM-YYYY),
# otherwise leave as None.
consolidated_output = None
if consolidated_dfs:
    try:
        # vlan_report['DATE'] should be text in DD-MMM-YYYY format; parse and find the max
        parsed = pd.to_datetime(vlan_report['DATE'], format='%d-%b-%Y', errors='coerce')
        if parsed.notna().any():
            consolidated_date = parsed.max().strftime('%d-%b-%Y')
        else:
            consolidated_date = output_date
    except Exception:
        consolidated_date = output_date
    consolidated_output = f'/home/asghar_attari1211/RnD/Consolidated_VLAN_Report_{consolidated_date}.xlsx'

def parse_service(text):
    if '_2G' in text or ' RAN_2G' in text or '2G' in text:
        return '2G'
    if '_3G' in text or '3G' in text:
        return '3G'
    if '_4G' in text or '4G' in text:
        return '4G'
    if '_5G' in text or '5G' in text:
        return '5G'
    return 'UNKNOWN'


def normalize_site(site):
    return str(site).strip().upper() if pd.notna(site) else None


def print_progress(message, verbose):
    if verbose:
        print(message)


def find_column_like(df, keywords):
    """Find first column in df whose name contains any of the keywords (case-insensitive).
    Returns the original column name or None if not found."""
    for col in df.columns:
        lcol = str(col).lower()
        for kw in keywords:
            if kw.lower() in lcol:
                return col
    return None


def site_role(site, phys_graph, ofn_types, fttt_sites, vsat_sites):
    if site in ofn_types:
        return 'OFN'
    if site in vsat_sites:
        return 'VSAT'
    if site in fttt_sites:
        neighbors = phys_graph.get(site, set())
        return 'FTTT-Tail' if len(neighbors) == 1 else 'FTTT-Hub'
    neighbors = phys_graph.get(site, set())
    if len(neighbors) == 1:
        return 'Tail'
    if len(neighbors) == 0:
        return 'Isolated'
    return 'Hub'


def tail_hub_status(role):
    if role in {'Tail', 'FTTT-Tail'}:
        return 'Tail'
    if role in {'Hub', 'FTTT-Hub'}:
        return 'Hub'
    if role == 'OFN':
        return 'OFN'
    return str(role)


def find_path_to_ofn(start_site, phys_graph, ofn_types, visited=None):
    if visited is None:
        visited = set()
    if start_site in visited:
        return None
    visited.add(start_site)
    if start_site in ofn_types:
        return [start_site]
    for nbr in phys_graph.get(start_site, set()):
        if nbr not in visited:
            path = find_path_to_ofn(nbr, phys_graph, ofn_types, visited.copy())
            if path:
                return [start_site] + path
    return None


def choose_best_candidate(candidates, vlan, service, route_library, ofn_types):
    if not candidates:
        return None
    # 1) Prefer candidates that are themselves Ring OFNs
    ring_ofns = [c for c in candidates if c in ofn_types and ofn_types[c].get('status') == 'Ring']
    if ring_ofns:
        return sorted(ring_ofns)[0]

    # 2) Prefer candidates whose known route terminates at a Ring OFN
    ring_via_candidates = []
    for c in candidates:
        r = route_library.get((c, vlan, service))
        if r and r.get('path') and r['path'][-1] in ofn_types and ofn_types.get(r['path'][-1], {}).get('status') == 'Ring':
            ring_via_candidates.append(c)
    if ring_via_candidates:
        return sorted(ring_via_candidates)[0]

    # 3) Prefer candidates that have a completed route
    completed = [c for c in candidates if (c, vlan, service) in route_library and route_library[(c, vlan, service)]['status'].startswith('Complete')]
    if completed:
        return sorted(completed)[0]

    # 4) Fallback to alphabetical deterministic choice
    return sorted(candidates)[0]


def is_spur_ofn(site, ofn_types):
    if site not in ofn_types:
        return False
    return ofn_types[site]['status'] != 'Ring'


def calculate_distance(coord1, coord2):
    """Calculate simple Euclidean distance between two coordinates (lat, lon)"""
    if not coord1 or not coord2:
        return float('inf')
    try:
        lat1, lon1 = float(coord1[0]), float(coord1[1])
        lat2, lon2 = float(coord2[0]), float(coord2[1])
        # Simple Euclidean distance (not geodesic, but good enough for comparison)
        return ((lat2 - lat1)**2 + (lon2 - lon1)**2)**0.5
    except (TypeError, ValueError):
        return float('inf')


def get_ofn_neighbors(site, phys_graph, ofn_types):
    """Get all OFN neighbors of a site from the physical graph"""
    neighbors = phys_graph.get(site, set())
    ofn_neighbors = [nbr for nbr in neighbors if nbr in ofn_types]
    return ofn_neighbors


def classify_path(path, service, ofn_types, fttt_sites, start_site, route_status=None):
    if route_status == 'Complete-Spur':
        return 'Complete'
    drop = path[-1]
    if drop in ofn_types:
        return 'Complete'
    if start_site in fttt_sites and service in {'3G', '4G', '5G'}:
        return 'Complete'
    if len(path) == 1 and start_site in fttt_sites:
        return 'Complete'
    if len(path) > 1:
        return 'Incomplete'
    return 'Broken'


vlan_report.columns = [str(c).strip() for c in vlan_report.columns]
consolidated_report_date = str(vlan_report['DATE'].dropna().iloc[0]).split(' ')[0] if not vlan_report.empty else 'unknown_date'
if skipped_files:
    print('⚠️ The following VLAN report files were skipped due to missing date in filename:')
    for sf in skipped_files:
        print(' -', sf)

vlan_presence_map = defaultdict(set)
direct_nms = defaultdict(set)
reverse_nms = defaultdict(set)
evidence_by_far_end = defaultdict(set)
all_nms_sites = set()

for _, row in vlan_report.iterrows():
    vid = str(row['VLAN ID']).strip()
    if not vid or vid.lower() in {'nan', 'none'}:
        continue

    # NOTE: Treat `VLAN Service Name` as the Near-End and `NE NAME` as the Far-End
    svc_sites = extract_sites(row.get('VLAN Service Name'))
    ne_sites = extract_sites(row.get('NE NAME'))
    # require at least one Near-End (VLAN Service Name) to proceed
    if not svc_sites:
        continue

    near = svc_sites[0]
    all_nms_sites.add(near)
    vlan_presence_map[vid].add(near)

    # register Far-End(s) coming from NE NAME column
    for far_end in ne_sites:
        if far_end == near:
            continue
        all_nms_sites.add(far_end)
        vlan_presence_map[vid].add(far_end)
        # direct_nms maps (NearEnd, VLAN) -> FarEnd(s)
        direct_nms[(near, vid)].add(far_end)
        # reverse_nms maps (FarEnd, VLAN) -> NearEnd(s)
        reverse_nms[(far_end, vid)].add(near)
        evidence_by_far_end[(far_end, vid)].add(near)
        evidence_by_far_end[(near, vid)].add(far_end)


print('🚀 [2/10] Building physical neighbor graph...')
lr_sheet = pd.read_excel(FILE_PATH, sheet_name='LR')
phys_graph = defaultdict(set)
for _, row in lr_sheet.iterrows():
    s1_sites = extract_sites(row.get('Site name S1'))
    s2_sites = extract_sites(row.get('Site name S2'))
    if not s1_sites or not s2_sites:
        continue
    s1, s2 = s1_sites[0], s2_sites[0]
    phys_graph[s1].add(s2)
    phys_graph[s2].add(s1)

# Identify LR column names for partner lookup
s1_col = find_column_like(lr_sheet, ['Site name S1', 'Site S1', 'S1'])
s2_col = find_column_like(lr_sheet, ['Site name S2', 'Site S2', 'S2'])


def find_lr_partner(site):
    """Return the LR partner site for `site` using the LR sheet S1/S2 columns.
    Returns normalized partner site or None."""
    try:
        if s1_col and s2_col:
            # search S1 column
            mask1 = lr_sheet[s1_col].apply(lambda x: normalize_site(extract_sites(x)[0]) == site if pd.notna(x) and extract_sites(x) else False)
            sub1 = lr_sheet[mask1]
            if not sub1.empty:
                val = sub1.iloc[0][s2_col]
                partners = extract_sites(val)
                return normalize_site(partners[0]) if partners else None

            # search S2 column
            mask2 = lr_sheet[s2_col].apply(lambda x: normalize_site(extract_sites(x)[0]) == site if pd.notna(x) and extract_sites(x) else False)
            sub2 = lr_sheet[mask2]
            if not sub2.empty:
                val = sub2.iloc[0][s1_col]
                partners = extract_sites(val)
                return normalize_site(partners[0]) if partners else None
    except Exception:
        return None
    return None


print('🚀 [3/10] Loading termination lists and SL sites...')
ofn_df = pd.read_excel(FILE_PATH, sheet_name='OFN')
ofn_types = {}
# flexible column detection for OFN sheet
ofn_id_col = find_column_like(ofn_df, ['OFN ID', 'ID', 'Site ID', 'SITE'])
ofn_status_col = find_column_like(ofn_df, ['OFN Status', 'Status', 'State'])
ofn_coloc_col = find_column_like(ofn_df, ['Colocation', 'Co-location', 'Colocated'])
for _, row in ofn_df.iterrows():
    site = normalize_site(row.get(ofn_id_col) if ofn_id_col else row.get('OFN ID'))
    status = str(row.get(ofn_status_col, '')).strip().title() if pd.notna(row.get(ofn_status_col, '')) else ''
    colocation = str(row.get(ofn_coloc_col, '')).strip().title() if pd.notna(row.get(ofn_coloc_col, '')) else ''
    if site:
        ofn_types[site] = {'status': status, 'colocation': colocation}

# OFN NODES RULE: Add missing OFN sites (sites that appear in VLAN reports but are missing from OFN sheet)
missing_ofn_sites = ['KHB108', 'KTM102']  # Sites missing in report but should be treated as OFNs
for missing_ofn in missing_ofn_sites:
    if missing_ofn not in ofn_types:
        ofn_types[missing_ofn] = {'status': 'Ring (Built)', 'colocation': False}
        print_progress(f'ℹ️ OFN NODES RULE: Added missing OFN site {missing_ofn} as Ring (Built)', args.verbose)

fttt_df = pd.read_excel(FILE_PATH, sheet_name='FTTT')
fttt_id_col = find_column_like(fttt_df, ['FTTT ID', 'ID', 'Site ID', 'SITE'])
if fttt_id_col:
    fttt_sites = {normalize_site(x) for x in fttt_df[fttt_id_col] if pd.notna(x)}
else:
    # fallback to any plausible first column
    fttt_sites = {normalize_site(x) for x in fttt_df.iloc[:, 0] if pd.notna(x)}

vsat_df = pd.read_excel(FILE_PATH, sheet_name='VSAT')
if 'Site Name' in vsat_df.columns and 'Type' in vsat_df.columns:
    vsat_sites = {normalize_site(x): str(y).strip() for x, y in zip(vsat_df['Site Name'], vsat_df['Type']) if pd.notna(x)}
else:
    vsat_id_col = find_column_like(vsat_df, ['VSAT ID', 'VSATID', 'Site ID', 'SITE', 'ID'])
    vsat_type_col = find_column_like(vsat_df, ['Type', 'VSAT Type', 'Class'])
    if vsat_id_col and vsat_type_col:
        vsat_sites = {normalize_site(x): str(y).strip() for x, y in zip(vsat_df[vsat_id_col], vsat_df[vsat_type_col]) if pd.notna(x)}
    else:
        print(f"⚠️ VSAT sheet columns not found. Available columns: {list(vsat_df.columns)}")
        vsat_sites = {}

try:
    sites_df = pd.read_excel(FILE_PATH, sheet_name='Sites')
except Exception:
    # Try to find any sheet that looks like a sites table (has lat/long columns)
    xl_all = pd.ExcelFile(FILE_PATH)
    sites_df = pd.DataFrame()
    for sn in xl_all.sheet_names:
        try:
            tmp = pd.read_excel(FILE_PATH, sheet_name=sn, nrows=5)
        except Exception:
            continue
        lcols = [str(c).lower() for c in tmp.columns]
        if any('lat' in c for c in lcols) and any('lon' in c or 'long' in c for c in lcols):
            sites_df = pd.read_excel(FILE_PATH, sheet_name=sn)
            break

site_id_col = find_column_like(sites_df, ['Site ID', 'ID', 'SITE', 'Site Name'])
lat_col = find_column_like(sites_df, ['Lat', 'Latitude'])
long_col = find_column_like(sites_df, ['Long', 'Longitude', 'Lon'])
if not sites_df.empty and site_id_col and lat_col and long_col:
    site_coords = {normalize_site(x): (y, z) for x, y, z in zip(sites_df[site_id_col], sites_df[lat_col], sites_df[long_col]) if pd.notna(x)}
else:
    site_coords = {}

# Read SL sheet with headers to get display names and coordinates
try:
    sl_df = pd.read_excel(FILE_PATH, sheet_name='SL')
    if 'Site Name' in sl_df.columns:
        sl_labels = {normalize_site(r): str(n) for r, n in zip(sl_df['Site Name'].apply(lambda x: extract_sites(x)[0] if extract_sites(x) else x), sl_df['Site Name']) if pd.notna(r)}
    else:
        # fallback to first column
        sl_df = pd.read_excel(FILE_PATH, sheet_name='SL', header=None)
        sl_labels = {normalize_site(x): x for x in sl_df.iloc[:, 0] if pd.notna(x)}
    # Add coordinates from SL sheet when available
    if 'Latitude' in sl_df.columns and 'Longitude' in sl_df.columns and 'Site Name' in sl_df.columns:
        for _, r in sl_df.iterrows():
            try:
                site_id_candidates = extract_sites(r['Site Name'])
                if not site_id_candidates:
                    continue
                sid = normalize_site(site_id_candidates[0])
                lat = r['Latitude']
                lon = r['Longitude']
                if pd.notna(lat) and pd.notna(lon):
                    site_coords.setdefault(sid, (lat, lon))
            except Exception:
                continue
except Exception:
    sl_labels = {}

sl_sheet = pd.read_excel(FILE_PATH, sheet_name='SL', header=None)
sl_sites = [normalize_site(x) for x in sl_sheet.iloc[:, 0] if pd.notna(x)]
sl_sites = [s for s in sl_sites if s != 'SITE NAME']
sl_sites = list(dict.fromkeys(sl_sites))

# Ensure OFN sites are present in SL processing list so they are included in Routes (Execution Prioritization P0)
for _ofn in ofn_types.keys():
    if _ofn not in sl_sites:
        sl_sites.append(_ofn)
        print_progress(f'ℹ️ Execution Prioritization: added OFN site {_ofn} to SL processing', args.verbose)

# If user supplied --filter-sites, also restrict SL processing list to matching sites
if args.filter_sites:
    prefixes = [p.strip().upper() for p in args.filter_sites.split(',')]
    sl_sites = [s for s in sl_sites if any(s.startswith(pref) for pref in prefixes)]
    print_progress(f'ℹ️ SL processing restricted to sites starting with: {prefixes} (count={len(sl_sites)})', args.verbose)

vlan_db = pd.read_excel(FILE_PATH, sheet_name='VLAN DB')
vlan_db['Site_Clean'] = vlan_db['Site'].apply(lambda x: extract_sites(x)[0] if extract_sites(x) else None)
vlan_db['Site_Clean'] = vlan_db['Site_Clean'].apply(normalize_site)
vlan_db['Service'] = vlan_db['Site'].apply(parse_service)
vlan_db = vlan_db[vlan_db['Site_Clean'].notna()]

if args.filter_sites:
    prefixes = [p.strip().upper() for p in args.filter_sites.split(',')]
    vlan_db = vlan_db[vlan_db['Site_Clean'].str.startswith(tuple(prefixes))]
    print_progress(f'ℹ️ Filtering VLAN DB to sites starting with: {prefixes}', args.verbose)

if args.sample_sites > 0:
    sl_sites = sl_sites[:args.sample_sites]
    print_progress(f'ℹ️ Sample mode active: limiting SL processing to {len(sl_sites)} sites', args.verbose)

if args.sample_rows > 0:
    vlan_db = vlan_db.head(args.sample_rows)
    print_progress(f'ℹ️ Sample mode active: limiting VLAN DB to {len(vlan_db)} rows', args.verbose)

print_progress(f'ℹ️ Final SL site count: {len(sl_sites)}; VLAN DB row count: {len(vlan_db)}', args.verbose)
print('🚀 [4/10] Preparing mismatch diagnostics...')
mismatch_records = []
sl_set = set(sl_sites)
vlan_db_sites = set(vlan_db['Site_Clean'].unique())
for site in sorted(set(list(ofn_types.keys()) + list(fttt_sites))):
    if site not in sl_set:
        mismatch_records.append({'Source': 'OFN/FTTT', 'Site ID': site, 'Reason': 'Site appears in OFN/FTTT lists but is missing from SL'})

for site in sorted(sl_set):
    if site not in vlan_db_sites:
        mismatch_records.append({'Source': 'SL', 'Site ID': site, 'Reason': 'Site appears in SL but is missing from VLAN DB'})
    if site not in all_nms_sites:
        mismatch_records.append({'Source': 'SL', 'Site ID': site, 'Reason': 'Site appears in SL but is missing from VLAN Report'})

for site in sorted(all_nms_sites):
    # Do not flag VLAN-report sites that are present in SL — SL sites are
    # not required to be OFN/FTTT/VSAT and should not be considered a discrepancy.
    if site not in ofn_types and site not in fttt_sites and site not in sl_set:
        mismatch_records.append({'Source': 'VLAN Report', 'Site ID': site, 'Reason': 'Site appears in VLAN Report but is not in OFN/FTTT lists'})

mismatch_df = pd.DataFrame(mismatch_records).drop_duplicates().sort_values(['Source', 'Site ID'])


print('🚀 [5/10] Seeding route library with OFN and FTTT terminals...')
route_library = {}
library_state = []
logic_verification = []


def store_route(site, vlan, service, path, status, stage, comment='', hop_method='', priority=''):
    key = (site, vlan, service)
    route_library[key] = {
        'path': path,
        'status': status,
        'stage': stage,
        'comment': comment,
        'hop_method': hop_method,
        'priority': priority,
    }
    library_state.append({
        'Stage': stage,
        'Site ID': site,
        'Service': service,
        'VLAN': vlan,
        'Path': ' > '.join(path),
        'Status': status,
        'Comment': comment,
        'Priority': priority,
    })


for ofn_site, ofn_info in ofn_types.items():
    # OFN nodes are drop points for all services (2G/3G/4G/5G).
    services = ['2G', '3G', '4G', '5G']
    for svc in services:
        store_route(ofn_site, 'OFN-SEED', svc, [ofn_site], 'Complete', 'OFN Seed', f'OFN termination stored ({ofn_info["status"]}, {"LM" if ofn_info["colocation"] else "Built"})', 'OFN-SEED', '0')

for _, row in vlan_db.iterrows():
    site = row['Site_Clean']
    service = row['Service']
    vlan = str(row['VLAN']).strip()
    if site in fttt_sites and service in {'3G', '4G', '5G'}:
        key = (site, vlan, service)
        if key not in route_library:
            store_route(site, vlan, service, [site], 'Complete', 'FTTT Seed', 'FTTT self-terminating route', 'FTTT-SEED', '0')

for vsat_site, vsat_type in vsat_sites.items():
    store_route(vsat_site, 'VSAT-SEED', 'ALL', [vsat_site], 'Complete', 'VSAT Seed', f'VSAT {vsat_type} termination stored', 'VSAT-SEED', '0')


print('🚀 [6/10] Processing SL sites in tail-first order...')
ordered_sites = [(site, site_role(site, phys_graph, ofn_types, fttt_sites, vsat_sites)) for site in sl_sites]
# EXECUTION PRIORITIZATION RULE (P0-P4):
# Desired sequence: OFN, VSAT, FTTT Tail, MW Tail, FTTT HUB, MW HUB, Isolated
priority_map = {
    'OFN': 0,
    'FTTT-Tail': 1,
    'VSAT': 2,
    'Tail': 3,
    # P4: All MW and FTTT HUB Sites should be same priority
    'FTTT-Hub': 4,
    'Hub': 4,
    'Isolated': 6,
}
ordered_sites.sort(key=lambda x: (priority_map.get(x[1], 9), x[0]))
print_progress(f'ℹ️ Ordered sites to process: {len(ordered_sites)} (P0-P4 prioritization applied)', args.verbose)


def trace_route(current_site, vlan, service, is_start_fttt, visited=None, origin=None, first_ofn=None, forced_lr_depth=0):
    if visited is None:
        visited = []
    if origin is None:
        origin = current_site

    if current_site in visited:
        return [current_site]
    visited.append(current_site)

    stored_key = (current_site, vlan, service)
    stored = route_library.get(stored_key)
    if stored and stored['status'].startswith('Complete'):
        if DEBUG_SITE == origin:
            print(f'🔍 {origin}: stopping at cached complete route for {current_site} -> {stored["path"]}')
        return stored['path']

    # Topology flags for this node
    is_ofn = current_site in ofn_types
    is_ring = ofn_types.get(current_site, {}).get('status') == 'Ring'
    is_spur = is_spur_ofn(current_site, ofn_types)
    is_fttt = current_site in fttt_sites

    # Termination rules:
    # OFN NODES RULE: If this node is an OFN, it is the drop node for ALL services
    # (2G and Non-2G) per Rule-02 — no VLAN tracing required.
    if is_ofn:
        return [current_site]

    # FTTT SITES RULE: All FTTT nodes are themselves drop nodes for their Non-2G services.
    # For 2G, strict VLAN-based tracing applies (Near-End then Far-End logic below).
    if is_fttt and service != '2G':
        return [current_site]

    # - For FTTT hubs: non-2G services terminate at the FTTT start when tracing starts there.
    if service != '2G' and is_fttt and is_start_fttt:
        return [current_site]

    physical_neighbors = phys_graph.get(current_site, set())
    next_hop = None
    hop_method = ''
    priority = ''
    debug_active = DEBUG_SITE == origin
    # Rule 04(a): If this is a Microwave Tail (not FTTT Tail) with NO VLAN trace
    # anywhere in the VLAN reports, infer its LR partner as the next hop.
    # This applies only to non-FTTT, non-OFN, non-VSAT tail sites.
    try:
        current_role = site_role(current_site, phys_graph, ofn_types, fttt_sites, vsat_sites)
    except Exception:
        current_role = site_role(current_site, phys_graph, ofn_types, fttt_sites, vsat_sites)
    if current_role == 'Tail' and current_site not in fttt_sites and current_site not in ofn_types and current_site not in vsat_sites:
        # no VLAN evidence anywhere for this site
        if current_site not in all_nms_sites:
            lr_partner = find_lr_partner(current_site)
            if lr_partner and lr_partner not in [] and lr_partner not in []:
                # only set next_hop here when we truly have no VLAN evidence
                if lr_partner not in []:
                    next_hop = lr_partner
                    hop_method = 'LR-ONLY'
                    priority = 'LR0'
                    if DEBUG_SITE == origin:
                        print(f'🔍 TRACE {origin}: Rule04 LR-only applied for tail {current_site} -> {lr_partner}')
    if debug_active:
        print(f'🔍 TRACE {origin}: at {current_site}, VLAN={vlan}, Service={service}, visited={visited}, neighbors={sorted(physical_neighbors)}')

    near_candidates = [nbr for nbr in reverse_nms.get((current_site, vlan), set()) if nbr in physical_neighbors and nbr not in visited]
    far_candidates = [nbr for nbr in direct_nms.get((current_site, vlan), set()) if nbr in physical_neighbors and nbr not in visited]

    if debug_active:
        print(f'🔍 TRACE {origin}: near_candidates={sorted(near_candidates)}, far_candidates={sorted(far_candidates)}')

    # Candidate selection already prefers near_candidates before far_candidates,
    # which aligns with the strict rule: prefer exact Near-End VLAN hits, then Far-End.
    if not next_hop:
        next_hop = choose_best_candidate(near_candidates, vlan, service, route_library, ofn_types)
        if next_hop:
            hop_method = 'NMS-DIR'
            priority = '1'
    if not next_hop:
        next_hop = choose_best_candidate(far_candidates, vlan, service, route_library, ofn_types)
        if next_hop:
            hop_method = 'NMS-DIR'
            priority = '2'

    if not next_hop and not (service == '2G' and is_spur):
        candidate_neighbors = [nbr for nbr in physical_neighbors if nbr not in visited and nbr in vlan_presence_map.get(vlan, set())]
        if debug_active:
            print(f'🔍 TRACE {origin}: vlan_presence_candidates={sorted(candidate_neighbors)}')
        next_hop = choose_best_candidate(candidate_neighbors, vlan, service, route_library, ofn_types)
        if next_hop:
            hop_method = 'VLAN-PRESENCE'
            priority = '3'

    if not next_hop and len(physical_neighbors) == 1 and not (service == '2G' and is_spur):
        potential = next(iter(physical_neighbors))
        if potential not in visited:
            next_hop = potential
            hop_method = 'LR-TAIL'
            priority = '4'

    # SITE WITH 2+ OFN PHYSICAL CONNECTIVITY RULE:
    # If a site has physical connectivity with 2 or more OFNs:
    # - For Tail sites: Connect to a Ring OFN; if all are Spur, choose the physically closest
    # - For HUB sites: Check which child sites are connected to which OFN and follow that path
    current_role = site_role(current_site, phys_graph, ofn_types, fttt_sites, vsat_sites)
    ofn_neighbors = get_ofn_neighbors(current_site, phys_graph, ofn_types)
    
    if not next_hop and len(ofn_neighbors) >= 2:
        if current_role == 'Tail' or current_role == 'FTTT-Tail':
            # Tail site with 2+ OFNs: prefer Ring OFN, else closest Spur
            ring_ofns = [ofn for ofn in ofn_neighbors if ofn_types[ofn]['status'] == 'Ring' and ofn not in visited]
            if ring_ofns:
                next_hop = sorted(ring_ofns)[0]
            else:
                # All connecting OFNs are Spur; choose the physically closest
                current_coords = site_coords.get(current_site)
                closest_ofn = None
                min_distance = float('inf')
                for ofn in ofn_neighbors:
                    if ofn not in visited:
                        ofn_coords = site_coords.get(ofn)
                        distance = calculate_distance(current_coords, ofn_coords)
                        if distance < min_distance:
                            min_distance = distance
                            closest_ofn = ofn
                if closest_ofn:
                    next_hop = closest_ofn
            if next_hop:
                hop_method = '2OFN-TAIL'
                priority = '5a'
                if debug_active:
                    print(f'🔍 TRACE {origin}: 2+ OFN Tail rule applied, selected {next_hop}')
        
        elif current_role == 'Hub' or current_role == 'FTTT-Hub':
            # Hub site with 2+ OFNs: check which OFN children are connected to
            children = [nbr for nbr in physical_neighbors if nbr != current_site and site_role(nbr, phys_graph, ofn_types, fttt_sites, vsat_sites) in {'Tail', 'FTTT-Tail'}]
            ofn_by_children = {}
            for child in children:
                child_route = route_library.get((child, vlan, service))
                if child_route and child_route.get('path'):
                    for node in child_route['path']:
                        if node in ofn_neighbors:
                            if node not in ofn_by_children:
                                ofn_by_children[node] = []
                            ofn_by_children[node].append(child)
            
            if ofn_by_children:
                # Choose OFN with most children, breaking ties alphabetically
                sorted_ofns = sorted(ofn_by_children.items(), key=lambda x: (-len(x[1]), x[0]))
                next_hop = sorted_ofns[0][0]
                hop_method = '2OFN-HUB'
                priority = '5b'
                if debug_active:
                    print(f'🔍 TRACE {origin}: 2+ OFN Hub rule applied, selected {next_hop} for {len(sorted_ofns[0][1])} children')

    # HUB with 2+ connections, one to OFN (fallback if 2+ OFN rule not applicable)
    if not next_hop and site_role(current_site, phys_graph, ofn_types, fttt_sites, vsat_sites) == 'Hub' and len(physical_neighbors) >= 2:
        ofn_neighbors_unvisited = [nbr for nbr in physical_neighbors if nbr in ofn_types and nbr not in visited]
        if ofn_neighbors_unvisited and len(ofn_neighbors_unvisited) < 2:  # Only if not already handled by 2+ OFN rule
            valid_ofn = []
            for ofn in ofn_neighbors_unvisited:
                ofn_info = ofn_types[ofn]
                if not ofn_info['colocation'] or service in ['3G', '4G', '5G']:
                    valid_ofn.append(ofn)
            if valid_ofn:
                next_hop = sorted(valid_ofn)[0]
                hop_method = 'HUB-OFN'
                priority = '5'

    if debug_active:
        if next_hop:
            print(f'🔍 TRACE {origin}: selected next_hop={next_hop} hop_method={hop_method} priority={priority}')
        else:
            print(f'🔍 TRACE {origin}: no next hop found at {current_site}')

    if next_hop:
        path = [current_site] + trace_route(next_hop, vlan, service, is_start_fttt, visited.copy(), origin, first_ofn, forced_lr_depth)
        if any(node in ofn_types for node in path) or (is_fttt and service in {'3G', '4G', '5G'}):
            status = 'Complete'
        else:
            status = 'Incomplete'
        store_route(current_site, vlan, service, path, status, 'Trace', f'Followed {hop_method}', hop_method, priority)
        logic_verification.append({
            'Origin': origin,
            'Current': current_site,
            'Next Hop': next_hop,
            'VLAN': vlan,
            'Service': service,
            'Hop Method': hop_method,
            'Priority': priority,
            'Path': ' > '.join(path),
            'Status': status,
        })
        return path

    if is_spur and service == '2G':
        store_route(current_site, vlan, service, [current_site], 'Complete-Spur', '2G Spur Drop', 'Spur node terminates 2G when no ring evidence')
        logic_verification.append({
            'Origin': origin,
            'Current': current_site,
            'Next Hop': '',
            'VLAN': vlan,
            'Service': service,
            'Hop Method': 'SPUR-DROP',
            'Priority': '5',
            'Path': current_site,
            'Status': 'Complete-Spur',
        })
        return [current_site]

    # No next hop found. For 2G we may need to rollback to the first OFN seen earlier
    # No next hop found. Try alternate-service adoption for non-2G services
    if service != '2G':
        # prefer 3G then 5G then 4G as alternate fallbacks (skip current service)
        for alt in ['3G', '5G', '4G']:
            if alt == service:
                continue
            alt_key = (current_site, vlan, alt)
            alt_route = route_library.get(alt_key)
            if alt_route and alt_route['status'].startswith('Complete') and alt_route['path'][-1] in ofn_types:
                path = alt_route['path']
                store_route(current_site, vlan, service, path, 'Incomplete', 'Followed-Alt-Service', f'Followed {alt} Route', 'FOLLOWED-ALT', '9')
                logic_verification.append({
                    'Origin': origin,
                    'Current': current_site,
                    'Next Hop': path[1] if len(path) > 1 else '',
                    'VLAN': vlan,
                    'Service': service,
                    'Hop Method': f'FOLLOWED-{alt}',
                    'Priority': '9',
                    'Path': ' > '.join(path),
                    'Status': 'Incomplete',
                })
                return path

    # 2G/NON-2G OFN FALLBACK RULE:
    # If 2G service didn't reach an OFN but any Non-2G service did,
    # follow the Non-2G path up to the OFN as a fallback for 2G
    if service == '2G':
        # Check if any Non-2G service from same site/VLAN reached an OFN
        for alt_service in ['3G', '4G', '5G']:
            alt_key = (current_site, vlan, alt_service)
            alt_route = route_library.get(alt_key)
            if alt_route and alt_route['status'].startswith('Complete') and alt_route['path'][-1] in ofn_types:
                path = alt_route['path']
                store_route(current_site, vlan, service, path, 'Incomplete', '2G-Follows-Non2G', f'2G followed {alt_service} route to OFN', 'FOLLOWED-NON2G', '9')
                logic_verification.append({
                    'Origin': origin,
                    'Current': current_site,
                    'Next Hop': path[1] if len(path) > 1 else '',
                    'VLAN': vlan,
                    'Service': service,
                    'Hop Method': f'FOLLOWED-{alt_service}',
                    'Priority': '9',
                    'Path': ' > '.join(path),
                    'Status': 'Incomplete',
                })
                if DEBUG_SITE == origin:
                    print(f'🔍 {origin}: 2G fallback to Non-2G ({alt_service}) path: {" > ".join(path)}')
                return path


    # Attempt LR auto-extension: if this site has exactly two physical neighbors
    # and one of them is already in the visited route, then try the other
    # neighbor as an automatic next hop and continue tracing. Limit recursive
    # LR-follow attempts to avoid runaway recursion; we allow up to 3 extensions.
    if len(physical_neighbors) == 2 and forced_lr_depth < 3:
        neighs = list(physical_neighbors)
        seen = [n for n in neighs if n in visited]
        other = [n for n in neighs if n not in visited]
        if seen and other:
            other_site = other[0]
            if debug_active:
                print(f'🔍 TRACE {origin}: LR-auto-extend from {current_site} to {other_site} (seen={seen}, depth={forced_lr_depth})')
            path = [current_site] + trace_route(other_site, vlan, service, is_start_fttt, visited.copy(), origin, first_ofn, forced_lr_depth + 1)
            # preserve incomplete status when no OFN reached
            status = 'Complete' if any(node in ofn_types for node in path) else 'Incomplete'
            store_route(current_site, vlan, service, path, status, 'LR-Auto-Extend', f'Auto-followed LR link to {other_site}', 'LR-AUTO', 'LR')
            logic_verification.append({
                'Origin': origin,
                'Current': current_site,
                'Next Hop': other_site,
                'VLAN': vlan,
                'Service': service,
                'Hop Method': 'LR-AUTO',
                'Priority': 'LR',
                'Path': ' > '.join(path),
                'Status': status,
            })
            return path

    store_route(current_site, vlan, service, [current_site], 'Broken', 'No next hop', f'No route from {current_site}')
    # Back engineering for broken routes
    ofn_path = find_path_to_ofn(current_site, phys_graph, ofn_types)
    if ofn_path and len(ofn_path) <= 3:  # certain if short path
        ofn_drop = ofn_path[-1]
        ofn_info = ofn_types[ofn_drop]
        if (not ofn_info['colocation'] and service == '2G') or (ofn_info['colocation'] and service in ['3G', '4G', '5G']):
            # Update the route
            route_library[(current_site, vlan, service)] = {
                'path': ofn_path,
                'status': 'Complete-BackEng',
                'stage': 'Back Engineering',
                'comment': f'Back engineered to {ofn_drop} via physical graph',
                'hop_method': 'BACK-ENG',
                'priority': '9',
            }
            library_state.append({
                'Stage': 'Back Engineering',
                'Site ID': current_site,
                'Service': service,
                'VLAN': vlan,
                'Path': ' > '.join(ofn_path),
                'Status': 'Complete-BackEng',
                'Comment': f'Back engineered to {ofn_drop}',
                'Priority': '9',
            })
            logic_verification.append({
                'Origin': origin,
                'Current': current_site,
                'Next Hop': ofn_path[1] if len(ofn_path) > 1 else '',
                'VLAN': vlan,
                'Service': service,
                'Hop Method': 'BACK-ENG',
                'Priority': '9',
                'Path': ' > '.join(ofn_path),
                'Status': 'Complete-BackEng',
            })
            return ofn_path
    # If this is a 2G trace and we previously encountered an OFN (`first_ofn`),
    # but downstream tracing didn't reach any OFN, roll the route back to the
    # `first_ofn` and treat that as the terminating point for 2G.
    if service == '2G' and first_ofn and first_ofn != current_site:
        if DEBUG_SITE == origin:
            print(f'🔁 {origin}: rolling back 2G route to first OFN {first_ofn} from {current_site}')
        # record rolled-back route for current node (higher stack frames will build full path)
        # Also record which sites were visited after the first OFN for auditing.
        try:
            if first_ofn in visited:
                idx = visited.index(first_ofn)
                visited_after = visited[idx+1:]
            else:
                visited_after = [current_site]
        except Exception:
            visited_after = [current_site]

        # Keep path oriented from current_site -> first_ofn for consistency with other routes
        rb_path = [current_site, first_ofn]
        route_library[(current_site, vlan, service)] = {
            'path': rb_path,
            'status': 'Complete-RolledBack',
            'stage': 'Rollback',
            'comment': f'Rolled back to first OFN {first_ofn}',
            'hop_method': 'ROLLBACK',
            'priority': 'RB',
            'visited_after_first_ofn': visited_after,
        }
        library_state.append({
            'Stage': 'Rollback',
            'Site ID': current_site,
            'Service': service,
            'VLAN': vlan,
            'Path': ' > '.join(rb_path),
            'Status': 'Complete-RolledBack',
            'Comment': f'Rolled back to first OFN {first_ofn}',
            'Visited After First OFN': ','.join(visited_after) if visited_after else '',
            'Priority': 'RB',
        })
        logic_verification.append({
            'Origin': origin,
            'Current': current_site,
            'Next Hop': first_ofn,
            'VLAN': vlan,
            'Service': service,
            'Hop Method': 'ROLLBACK',
            'Priority': 'RB',
            'Path': ' > '.join(rb_path + visited_after),
            'Status': 'Complete-RolledBack',
        })
        return rb_path
    logic_verification.append({
        'Origin': origin,
        'Current': current_site,
        'Next Hop': '',
        'VLAN': vlan,
        'Service': service,
        'Hop Method': 'NO-HOP',
        'Priority': '9',
        'Path': current_site,
        'Status': 'Broken',
    })
    return [current_site]


def infer_best_hub_neighbor(drop, service, vlan, route_library, ofn_types, fttt_sites, current_path=None):
    current_path = current_path or []
    neighbors = [nbr for nbr in phys_graph.get(drop, set()) if nbr != drop]
    if not neighbors:
        return None

    role_map = {nbr: site_role(nbr, phys_graph, ofn_types, fttt_sites, vsat_sites) for nbr in neighbors}
    hub_neighbors = [nbr for nbr, role in role_map.items() if role in {'Hub', 'FTTT-Hub'}]
    tail_neighbors = [nbr for nbr, role in role_map.items() if role in {'Tail', 'FTTT-Tail'}]
    debug_active = DEBUG_SITE and (DEBUG_SITE == drop or (current_path and DEBUG_SITE == current_path[0]))
    if debug_active:
        print(f'🔍 TOPOLOGY {DEBUG_SITE}: drop={drop}, neighbors={sorted(neighbors)}, role_map={role_map}, current_path={current_path}')

    if len(hub_neighbors) == 1 and len(tail_neighbors) == len(neighbors) - 1:
        if debug_active:
            print(f'🔍 TOPOLOGY {DEBUG_SITE}: star topology detected, selecting hub {hub_neighbors[0]}')
        return hub_neighbors[0], 'TOPOLOGY-HUB'

    if len(neighbors) == 2 and len(tail_neighbors) == 1 and len(hub_neighbors) == 1:
        if debug_active:
            print(f'🔍 TOPOLOGY {DEBUG_SITE}: chain topology detected, selecting hub {hub_neighbors[0]}')
        return hub_neighbors[0], 'TOPOLOGY-CHAIN'

    if len(neighbors) == 2 and len(hub_neighbors) == 2:
        remaining_hubs = [nbr for nbr in hub_neighbors if nbr not in current_path]
        if remaining_hubs:
            selected = sorted(remaining_hubs)[0]
            if debug_active:
                print(f'🔍 TOPOLOGY {DEBUG_SITE}: both neighbors are hubs, selecting remaining hub {selected}')
            return selected, 'TOPOLOGY-ALT-HUB'

        hub_candidates = []
        for hub in hub_neighbors:
            hub_route = route_library.get((hub, vlan, service))
            if hub_route and hub_route['status'].startswith('Complete') and hub_route['path'][-1] in ofn_types:
                hub_candidates.append((0, hub))
                continue

            tail_complete = []
            for nbr in phys_graph.get(hub, set()):
                if nbr == drop:
                    continue
                if site_role(nbr, phys_graph, ofn_types, fttt_sites, vsat_sites) in {'Tail', 'FTTT-Tail'}:
                    nbr_route = route_library.get((nbr, vlan, service))
                    if nbr_route and nbr_route['status'].startswith('Complete') and nbr_route['path'][-1] in ofn_types:
                        tail_complete.append(nbr)
            if tail_complete:
                hub_candidates.append((1, hub))

        if hub_candidates:
            selected = sorted(hub_candidates)[0][1]
            if debug_active:
                print(f'🔍 TOPOLOGY {DEBUG_SITE}: hub route evidence selected {selected}')
            return selected, 'TOPOLOGY-ROUTE-EVIDENCE'

    direct_complete = []
    for nbr in neighbors:
        neighbor_route = route_library.get((nbr, vlan, service))
        if neighbor_route and neighbor_route['status'].startswith('Complete') and neighbor_route['path'][-1] in ofn_types:
            direct_complete.append(nbr)
    if direct_complete:
        direct_complete = [nbr for nbr in direct_complete if nbr not in current_path] or direct_complete
        if debug_active:
            print(f'🔍 TOPOLOGY {DEBUG_SITE}: direct complete route chosen {sorted(direct_complete)[0]}')
        return sorted(direct_complete)[0], 'TOPOLOGY-DIRECT'

    service_order = ['4G', '3G', '2G']
    for alt_service in service_order:
        if alt_service == service:
            continue
        for nbr in hub_neighbors + tail_neighbors:
            if nbr in current_path:
                continue
            alt_route = route_library.get((nbr, vlan, alt_service))
            if alt_route and alt_route['status'].startswith('Complete'):
                if debug_active:
                    print(f'🔍 TOPOLOGY {DEBUG_SITE}: alternate service {alt_service} chosen via {nbr}')
                return nbr, f'TOPOLOGY-SERVICE-{alt_service}'

    for hub in hub_neighbors:
        logical_neighbors = evidence_by_far_end.get((hub, vlan), set())
        for candidate in logical_neighbors:
            if candidate in current_path or candidate == hub:
                continue
            candidate_route = route_library.get((candidate, vlan, service))
            if candidate_route and candidate_route['status'].startswith('Complete') and candidate_route['path'][-1] in ofn_types:
                if debug_active:
                    print(f'🔍 TOPOLOGY {DEBUG_SITE}: logical site {candidate} on hub {hub} has OFN route')
                return hub, 'TOPOLOGY-LOGICAL'

    return None


def deduce_media(site, fttt_sites, ofn_types):
    if site in ofn_types:
        return 'OFN'
    if site in fttt_sites:
        return 'FTTT'
    return 'MW'


def clear_fe_hops(result):
    for key in list(result.keys()):
        if key.startswith('FE'):
            result.pop(key, None)


def update_final_result(site, vlan, service, path, status, comment, fttt_sites, ofn_types):
    for res in final_results:
        if res['Site ID'] == site and res['VLAN'] == vlan and res['Service'] == service:
            res['Drop Node'] = path[-1]
            res['Drop Node OFN Status'] = f"{ofn_types.get(path[-1], {}).get('status', 'Not in OFN Sheet')} ({'LM' if ofn_types.get(path[-1], {}).get('colocation') else 'Built'})"
            res['Media Type'] = deduce_media(site, fttt_sites, ofn_types)
            res['Status'] = status
            res['Comment'] = comment
            clear_fe_hops(res)
            for i, hop in enumerate(path[1:], start=1):
                res[f'FE{i}'] = hop
            return True
    return False


def site_has_ofn_drop(site):
    """Return True if any known route for `site` currently drops to an OFN.
    This is used to skip retrace iterations for sites that already have
    at least one service reaching an OFN by any logic."""
    try:
        for (s, v, svc), r in route_library.items():
            if s != site:
                continue
            p = r.get('path')
            if p and p[-1] in ofn_types:
                return True
    except Exception:
        pass
    return False


final_results = []
all_routes = []


def process_site_list(site_list, start_index=1):
    """Process a list of (site, role) tuples and populate final_results and all_routes."""
    idx = start_index
    total = len(site_list)
    for site, role in site_list:
        if args.verbose and idx % 200 == 0:
            print(f'⏳ Processing site {idx}/{total}: {site} ({role})')
        idx += 1

        site_rows = vlan_db[vlan_db['Site_Clean'] == site]
        if site_rows.empty:
            continue
        is_start_fttt = site in fttt_sites
        stage_name = 'Tail' if role == 'Tail' else ('FTTT-Tail' if role == 'FTTT-Tail' else ('FTTT-Hub' if role == 'FTTT-Hub' else ('Hub' if role == 'Hub' else 'Other')))

        for _, row in site_rows.iterrows():
            vlan = str(row['VLAN']).strip()
            service = row['Service']
            path = trace_route(site, vlan, service, is_start_fttt)
            drop = path[-1]
            route_status = route_library.get((site, vlan, service), {}).get('status')
            status = classify_path(path, service, ofn_types, fttt_sites, site, route_status)
            site_type = role
            comment = route_library.get((site, vlan, service), {}).get('comment', '')

            media_type = deduce_media(site, fttt_sites, ofn_types)
            result = {
                'Site ID': site,
                'Service': service,
                'VLAN': vlan,
                'Site Role': site_type,
                'Tail-HUB Status': tail_hub_status(role),
                'Site Type': site_type,
                'Media Type': media_type,
                'Drop Node': drop,
                'Drop Node OFN Status': f"{ofn_types.get(drop, {}).get('status', 'Not in OFN Sheet')} ({'LM' if ofn_types.get(drop, {}).get('colocation') else 'Built'})",
                'Status': status,
                'Comment': comment,
                'NE': site,
            }
            for i, hop in enumerate(path[1:], start=1):
                result[f'FE{i}'] = hop
            updated = update_final_result(site, vlan, service, path, status, comment, fttt_sites, ofn_types)
            if not updated:
                final_results.append(result)

            # update or append in all_routes
            found = False
            for j, rec in enumerate(all_routes):
                if rec[0] == site and rec[1] == vlan and rec[2] == service:
                    all_routes[j] = (site, vlan, service, path, status)
                    found = True
                    break
            if not found:
                all_routes.append((site, vlan, service, path, status))


# Build phase-1 list: OFN sites, FTTT-Tail, and MW Tail (Tail but not FTTT)
phase1 = []
phase2 = []
for site, role in ordered_sites:
    if site in ofn_types or role == 'FTTT-Tail' or (role == 'Tail' and site not in fttt_sites):
        phase1.append((site, role))
    else:
        phase2.append((site, role))

# Build master list of all route subjects (site, vlan, service) from VLAN DB
# This allows Phase1 retrace to attempt learning across the entire dataset
# (including sites not yet processed in phase2) so completed OFN routes can
# propagate earlier.
master_subjects = []
seen_subjects = set()
for _, r in vlan_db.iterrows():
    site = r['Site_Clean']
    if site not in sl_sites:
        continue
    vlan = str(r['VLAN']).strip()
    service = r['Service']
    key = (site, vlan, service)
    if key in seen_subjects:
        continue
    seen_subjects.add(key)
    master_subjects.append(key)

# Process phase1 first so routes that are most likely to complete run before retrace
process_site_list(phase1, start_index=1)
# After processing OFN / FTTT-Tail / MW Tail sites, run retrace iterations
print('🔁 [Phase1] Retracing incomplete routes learned from phase1 (cross-dataset)...')
max_retrace_iterations = 10
for retrace_iteration in range(max_retrace_iterations):
    changed = False
    for (site, vlan, service) in master_subjects:
        # If any service of this site already drops to an OFN, skip all subjects for that site
        if site_has_ofn_drop(site):
            continue
        # skip subjects already completed
        cur_status = route_library.get((site, vlan, service), {}).get('status', '')
        if cur_status.startswith('Complete'):
            continue
        is_start_fttt = site in fttt_sites
        # attempt to trace using current library state
        new_path = trace_route(site, vlan, service, is_start_fttt)
        new_route_status = route_library.get((site, vlan, service), {}).get('status')
        new_status = classify_path(new_path, service, ofn_types, fttt_sites, site, new_route_status)
        # find existing path/status if present
        prev_path = None
        prev_status = ''
        for idx, rec in enumerate(all_routes):
            if rec[0] == site and rec[1] == vlan and rec[2] == service:
                prev_path = rec[3]
                prev_status = rec[4]
                break

        if prev_path is None or new_path != prev_path or new_status != prev_status:
            changed = True
            comment = route_library.get((site, vlan, service), {}).get('comment', '')
            if new_status.startswith('Complete'):
                store_route(site, vlan, service, new_path, new_status, 'Retrace-Phase1', comment, 'RETRACE', 'R')
                logic_verification.append({
                    'Origin': site,
                    'Current': new_path[-1],
                    'Next Hop': '' if len(new_path) == 1 else new_path[1],
                    'VLAN': vlan,
                    'Service': service,
                    'Hop Method': 'RETRACE',
                    'Priority': 'R',
                    'Path': ' > '.join(new_path),
                    'Status': new_status,
                })
            # update/add all_routes record
            found = False
            for j, rec in enumerate(all_routes):
                if rec[0] == site and rec[1] == vlan and rec[2] == service:
                    all_routes[j] = (site, vlan, service, new_path, new_status)
                    found = True
                    break
            if not found:
                all_routes.append((site, vlan, service, new_path, new_status))
            # update final_results (append if missing)
            updated = update_final_result(site, vlan, service, new_path, new_status, comment, fttt_sites, ofn_types)
            if not updated:
                # create a result entry similar to process_site_list
                media_type = deduce_media(site, fttt_sites, ofn_types)
                role = site_role(site, phys_graph, ofn_types, fttt_sites, vsat_sites)
                result = {
                    'Site ID': site,
                    'Service': service,
                    'VLAN': vlan,
                    'Site Role': role,
                    'Tail-HUB Status': tail_hub_status(role),
                    'Site Type': role,
                    'Media Type': media_type,
                    'Drop Node': new_path[-1],
                    'Drop Node OFN Status': f"{ofn_types.get(new_path[-1], {}).get('status', 'Not in OFN Sheet')} ({'LM' if ofn_types.get(new_path[-1], {}).get('colocation') else 'Built'})",
                    'Status': new_status,
                    'Comment': comment,
                    'NE': site,
                }
                for i, hop in enumerate(new_path[1:], start=1):
                    result[f'FE{i}'] = hop
                final_results.append(result)
    if changed:
        print_progress(f'ℹ️ Phase1 Retrace pass {retrace_iteration + 1}/{max_retrace_iterations} updated incomplete routes', args.verbose)
    if not changed:
        break
# Then process remaining sites
process_site_list(phase2, start_index=len(phase1)+1)


print('🚀 [7/10] Propagating completed routes through incomplete links...')
max_retrace_iterations = 10
for retrace_iteration in range(max_retrace_iterations):
    changed = False
    for idx, (site, vlan, service, path, status) in enumerate(all_routes):
        # Skip routes already complete or where the site already has any service reaching an OFN
        if status.startswith('Complete'):
            continue
        if site_has_ofn_drop(site):
            continue
        
        is_start_fttt = site in fttt_sites
        new_path = trace_route(site, vlan, service, is_start_fttt)
        new_route_status = route_library.get((site, vlan, service), {}).get('status')
        new_status = classify_path(new_path, service, ofn_types, fttt_sites, site, new_route_status)
        
        if new_path != path or new_status != status:
            changed = True
            comment = route_library.get((site, vlan, service), {}).get('comment', '')
            if new_status.startswith('Complete'):
                store_route(site, vlan, service, new_path, new_status, 'Retrace', comment, 'RETRACE', 'R')
                logic_verification.append({
                    'Origin': site,
                    'Current': new_path[-1],
                    'Next Hop': '' if len(new_path) == 1 else new_path[1],
                    'VLAN': vlan,
                    'Service': service,
                    'Hop Method': 'RETRACE',
                    'Priority': 'R',
                    'Path': ' > '.join(new_path),
                    'Status': new_status,
                })
            all_routes[idx] = (site, vlan, service, new_path, new_status)
            update_final_result(site, vlan, service, new_path, new_status, comment, fttt_sites, ofn_types)
    
    if changed:
        print_progress(f'ℹ️ Retrace pass {retrace_iteration + 1}/{max_retrace_iterations} updated incomplete routes', args.verbose)
    if not changed:
        break


print('🚀 [8/10] Running forensic gap repair for broken routes...')
print_progress(f'ℹ️ Evaluating {len(all_routes)} routes for forensic repair', args.verbose)
for site, vlan, service, path, status in all_routes:
    if status.startswith('Complete'):
        continue
    drop = path[-1]
    physical_neighbors = phys_graph.get(drop, set())
    stitched = False

    for neighbor in sorted(physical_neighbors):
        evidence_subjects = evidence_by_far_end.get((neighbor, vlan), set())
        if site not in evidence_subjects:
            continue
        neighbor_key = (neighbor, vlan, service)
        neighbor_route = route_library.get(neighbor_key)
        if not neighbor_route or not neighbor_route['status'].startswith('Complete'):
            continue

        new_path = path[:-1] + neighbor_route['path']
        if new_path[-1] in ofn_types:
            stitched = True
            bridge_comment = f'FORENSIC-STITCH via {neighbor}'
            store_route(site, vlan, service, new_path, 'Complete-Bridged', 'Forensic Repair', bridge_comment, 'FORENSIC-STITCH', 'B')
            logic_verification.append({
                'Origin': site,
                'Current': drop,
                'Next Hop': neighbor,
                'VLAN': vlan,
                'Service': service,
                'Hop Method': 'FORENSIC-STITCH',
                'Priority': 'B',
                'Path': ' > '.join(new_path),
                'Status': 'Complete-Bridged',
            })
            break

    if stitched:
        for res in final_results:
            if res['Site ID'] == site and res['VLAN'] == vlan and res['Service'] == service:
                new_drop = route_library[(site, vlan, service)]['path'][-1]
                res['Drop Node'] = new_drop
                res['Drop Node OFN Status'] = f"{ofn_types.get(new_drop, {}).get('status', 'Not in OFN Sheet')} ({'LM' if ofn_types.get(new_drop, {}).get('colocation') else 'Built'})"
                res['Media Type'] = deduce_media(site, fttt_sites, ofn_types)
                res['Status'] = 'Complete-Bridged'
                res['Comment'] = bridge_comment
                for i, hop in enumerate(route_library[(site, vlan, service)]['path'][1:], start=1):
                    res[f'FE{i}'] = hop

# ------------------------------------------------------------------
# [Rule] Two-site physical LR extension for incomplete routes
# If the last site has exactly 2 physical neighbors and one neighbor
# is already present in the route, append the other neighbor automatically.
for idx, (site, vlan, service, path, status) in enumerate(all_routes):
    if status.startswith('Complete'):
        continue
    drop = path[-1]
    physical_neighbors = phys_graph.get(drop, set())
    if len(physical_neighbors) != 2:
        continue
    neighbors = sorted(physical_neighbors)
    already_in_route = [nbr for nbr in neighbors if nbr in path]
    if len(already_in_route) != 1:
        continue
    next_hop = [nbr for nbr in neighbors if nbr not in path][0]
    if next_hop in path:
        continue

    new_path = path.copy()
    new_path.append(next_hop)
    route_comment = 'Followed LR link to second neighbor'
    store_route(site, vlan, service, new_path, 'Incomplete', 'LR Link Extension', route_comment, 'FOLLOWED-LR', 'LR')
    logic_verification.append({
        'Origin': site,
        'Current': drop,
        'Next Hop': next_hop,
        'VLAN': vlan,
        'Service': service,
        'Hop Method': 'FOLLOWED-LR',
        'Priority': 'LR',
        'Path': ' > '.join(new_path),
        'Status': 'Incomplete',
    })
    for res in final_results:
        if res['Site ID'] == site and res['VLAN'] == vlan and res['Service'] == service:
            res['Drop Node'] = next_hop
            res['Drop Node OFN Status'] = f"{ofn_types.get(next_hop, {}).get('status', 'Not in OFN Sheet')} ({'LM' if ofn_types.get(next_hop, {}).get('colocation') else 'Built'})"
            res['Media Type'] = deduce_media(site, fttt_sites, ofn_types)
            res['Status'] = 'Incomplete'
            res['Comment'] = route_comment
            for i, hop in enumerate(new_path[1:], start=1):
                res[f'FE{i}'] = hop
    all_routes[idx] = (site, vlan, service, new_path, 'Incomplete')

print('🚀 [9/10] Applying topology repair for disconnected routes...')
# Only iterate over incomplete routes to avoid unnecessary traversal of completed entries
incomplete_indices = [i for i, (_, _, _, _, st) in enumerate(all_routes) if not st.startswith('Complete')]
print_progress(f'ℹ️ Topology repair will evaluate {len(incomplete_indices)} incomplete routes', args.verbose)
cnt = 0
for idx in incomplete_indices:
    site, vlan, service, path, status = all_routes[idx]
    # increment counter only for routes we actually evaluate
    cnt += 1
    if cnt % 1000 == 0:
        print(cnt)
    drop = path[-1]
    if drop in ofn_types:
        continue

    inference = infer_best_hub_neighbor(drop, service, vlan, route_library, ofn_types, fttt_sites, current_path=path)
    if not inference:
        continue

    next_hop, method = inference
    if next_hop in path:
        continue

    next_route = route_library.get((next_hop, vlan, service))
    if next_route and next_route['status'].startswith('Complete') and next_route['path'][-1] in ofn_types:
        new_path = path.copy()
        if new_path[-1] != next_hop:
            new_path.append(next_hop)
        new_path.extend(next_route['path'][1:])
    else:
        alt_complete = None
        for ((src, vlan_key, svc), route) in route_library.items():
            if src == next_hop and route['status'].startswith('Complete') and route['path'][-1] in ofn_types:
                alt_complete = route
                break
        if not alt_complete:
            continue
        new_path = path.copy()
        if new_path[-1] != next_hop:
            new_path.append(next_hop)
        new_path.extend(alt_complete['path'][1:])

    inferred_comment = f'INFERRED-{method} via {next_hop}'
    existing_comment = route_library.get((site, vlan, service), {}).get('comment', '')
    route_comment = existing_comment if existing_comment else inferred_comment
    inferred_status = 'Complete-Inferred' if new_path[-1] in ofn_types else 'Incomplete-Inferred'
    store_route(site, vlan, service, new_path, inferred_status, 'Topology Repair', route_comment, method, 'T')
    logic_verification.append({
        'Origin': site,
        'Current': drop,
        'Next Hop': next_hop,
        'VLAN': vlan,
        'Service': service,
        'Hop Method': method,
        'Priority': 'T',
        'Path': ' > '.join(new_path),
        'Status': inferred_status,
    })

    for res in final_results:
        if res['Site ID'] == site and res['VLAN'] == vlan and res['Service'] == service:
            new_drop = new_path[-1]
            res['Drop Node'] = new_drop
            res['Drop Node OFN Status'] = f"{ofn_types.get(new_drop, {}).get('status', 'Not in OFN Sheet')} ({'LM' if ofn_types.get(new_drop, {}).get('colocation') else 'Built'})"
            res['Media Type'] = deduce_media(site, fttt_sites, ofn_types)
            res['Status'] = inferred_status
            res['Comment'] = route_comment
            for i, hop in enumerate(new_path[1:], start=1):
                res[f'FE{i}'] = hop
    # ------------------------------------------------------------------
    # [Rule] Follow Non-2G Services That Reach OFN
    # If any non-2G service (3G/4G/5G) from the same site reaches an OFN,
    # other non-2G services that did NOT reach an OFN should follow the
    # route of that service. This provides a deterministic fallback so that
    #, e.g., a 4G route will use the 3G route when 3G already reaches OFN.
    # If none of the non-2G services reach an OFN but the site's 2G route
    # reaches an OFN, then use the 2G route up to the first OFN as the
    # fallback for non-2G services (status remains Incomplete).
    # The code below enacts these rules and updates `route_library`,
    # `library_state`, and `final_results` accordingly.
    # (Silent) apply Non-2G follow rules to optimize incomplete non-2G routes
    # Build quick lookup of routes by (site, service) -> (vlan, path, status)
    site_service_map = defaultdict(list)
    for site, vlan, service, path, status in all_routes:
        site_service_map[site].append((service, vlan, path, status))

    # Preferred donor order when multiple non-2G donors exist
    donor_order = ['3G', '5G', '4G']
    changes = 0
    for site, svc_entries in site_service_map.items():
        # Do not apply follow rules to FTTT sites: non-2G services on FTTT
        # sites are self-terminating per Rule-03 and must not be overwritten.
        if site in fttt_sites:
            continue
        # collect donor services (non-2G) that reach OFN
        donors = {}
        for service, vlan, path, status in svc_entries:
            if service == '2G':
                continue
            route = route_library.get((site, vlan, service))
            if route and route['status'].startswith('Complete') and route['path'] and route['path'][-1] in ofn_types:
                donors[service] = (vlan, route['path'])

        # For each non-2G service that is incomplete or doesn't end in OFN, try to follow a donor
        for service, vlan, path, status in svc_entries:
            if service == '2G':
                continue
            cur_route = route_library.get((site, vlan, service))
            ends_in_ofn = cur_route and cur_route.get('path') and cur_route['path'][-1] in ofn_types
            if ends_in_ofn:
                continue

            # First preference: follow another non-2G donor route that reaches OFN
            selected_donor = None
            for d in donor_order:
                if d in donors:
                    selected_donor = (d, donors[d][0], donors[d][1])
                    break

            # If no non-2G donor found, consider 2G as fallback (use 2G up to first OFN)
            if not selected_donor:
                # find 2G route for this site
                tg = None
                for s2, v2, p2, st2 in svc_entries:
                    if s2 == '2G':
                        tg = (v2, p2, st2)
                        break
                if tg:
                    v2, p2, st2 = tg
                    # find first OFN in 2G path
                    first_ofn_idx = None
                    for i, node in enumerate(p2):
                        if node in ofn_types:
                            first_ofn_idx = i
                            break
                    if first_ofn_idx is not None and first_ofn_idx > 0:
                        # use subpath from site to first OFN inclusive
                        donor_path = p2[: first_ofn_idx + 1]
                        selected_donor = ('2G', v2, donor_path)

            if selected_donor:
                donor_service, donor_vlan, donor_path = selected_donor
                # Apply the followed route for this non-2G service
                # Keep status Incomplete (we followed another service's route)
                comment = f'Followed {donor_service} Route'
                route_library[(site, vlan, service)] = {
                    'path': donor_path,
                    'status': 'Incomplete',
                    'stage': 'Followed-Non2G' if donor_service != '2G' else 'Followed-2G',
                    'comment': comment,
                    'hop_method': f'FOLLOWED-{donor_service}',
                    'priority': 'F',
                }
                library_state.append({
                    'Stage': route_library[(site, vlan, service)]['stage'],
                    'Site ID': site,
                    'Service': service,
                    'VLAN': vlan,
                    'Path': ' > '.join(donor_path),
                    'Status': 'Incomplete',
                    'Comment': comment,
                    'Priority': 'F',
                })

                # Update final_results entry for this route
                for res in final_results:
                    if res['Site ID'] == site and res['VLAN'] == vlan and res['Service'] == service:
                        new_drop = donor_path[-1]
                        res['Drop Node'] = new_drop
                        res['Drop Node OFN Status'] = f"{ofn_types.get(new_drop, {}).get('status', 'Not in OFN Sheet')} ({'LM' if ofn_types.get(new_drop, {}).get('colocation') else 'Built'})"
                        res['Status'] = 'Incomplete'
                        res['Comment'] = comment
                        # clear existing FE* then set according to donor_path
                        clear_fe_hops(res)
                        for i, hop in enumerate(donor_path[1:], start=1):
                            res[f'FE{i}'] = hop
                changes += 1

    # completed Non-2G follow rules (silent)

print('🚀 [10/10] Exporting final workbook with audit sheets...')
try:
    print(f"DEBUG: df_final rows={len(final_results)}")
except Exception:
    pass
try:
    import psutil
    proc = psutil.Process()
    mem = proc.memory_info().rss
    print(f"DEBUG: process RSS memory: {mem/1024/1024:.1f} MB")
except Exception:
    try:
        print(f"DEBUG: ru_maxrss={resource.getrusage(resource.RUSAGE_SELF).ru_maxrss}")
    except Exception:
        pass
df_final = pd.DataFrame(final_results)
# ------------------------------------------------------------------
# Post-processing optimization pass on the DataFrame copy
# Apply Non-2G follow rules directly to `df_final` before export so
# the written workbook reflects adopted routes even in dry-run mode.
try:
    if not df_final.empty:
        # helper to extract full path list from a row (NE + FE1..FEn)
        def row_path(r):
            path = []
            ne = r.get('NE')
            if pd.notna(ne):
                path.append(ne)
            i = 1
            while True:
                col = f'FE{i}'
                if col not in r.index:
                    break
                v = r.get(col)
                if pd.isna(v) or v == '':
                    break
                path.append(v)
                i += 1
            return path

        donor_order = ['3G', '5G', '4G']
        updated = 0
        for site, grp in df_final.groupby('Site ID'):
            # Do not apply dataframe-level Non-2G follow rules to FTTT sites.
            if site in fttt_sites:
                continue
            # collect non-2G donors that reach OFN
            donors = {}
            two_g_path = None
            for _, r in grp.iterrows():
                svc = r['Service']
                path = row_path(r)
                drop = path[-1] if path else None
                if svc == '2G' and drop in ofn_types:
                    two_g_path = path
                if svc != '2G' and drop in ofn_types:
                    donors[svc] = path

            for idx, r in grp.iterrows():
                svc = r['Service']
                if svc == '2G':
                    continue
                path = row_path(r)
                drop = path[-1] if path else None
                if drop in ofn_types:
                    continue

                selected = None
                for d in donor_order:
                    if d in donors:
                        selected = ('NON2G', d, donors[d])
                        break
                if not selected and two_g_path:
                    # find first OFN in 2G path
                    first_ofn_idx = None
                    for i, node in enumerate(two_g_path):
                        if node in ofn_types:
                            first_ofn_idx = i
                            break
                    if first_ofn_idx is not None and first_ofn_idx > 0:
                        selected = ('2G', '2G', two_g_path[: first_ofn_idx + 1])

                if selected:
                    _, donor_svc, donor_path = selected
                    comment = f'Followed {donor_svc} Route'
                    # update df_final in-place
                    rowmask = (df_final['Site ID'] == site) & (df_final['Service'] == svc) & (df_final['VLAN'] == r['VLAN'])
                    df_final.loc[rowmask, 'Drop Node'] = donor_path[-1]
                    df_final.loc[rowmask, 'Drop Node OFN Status'] = f"{ofn_types.get(donor_path[-1],{}).get('status','Not in OFN Sheet')} ({'LM' if ofn_types.get(donor_path[-1],{}).get('colocation') else 'Built'})"
                    df_final.loc[rowmask, 'Status'] = 'Incomplete'
                    df_final.loc[rowmask, 'Comment'] = comment
                    # clear FE* for that row then set according to donor_path
                    max_fe = max([int(c[2:]) for c in df_final.columns if str(c).startswith('FE')] or [0])
                    for i in range(1, max_fe+1):
                        col = f'FE{i}'
                        if col in df_final.columns:
                            df_final.loc[rowmask, col] = ''
                    for i, hop in enumerate(donor_path[1:], start=1):
                        col = f'FE{i}'
                        if col not in df_final.columns:
                            df_final[col] = ''
                        df_final.loc[rowmask, f'FE{i}'] = hop
                    updated += 1
        if updated:
            # completed DataFrame-level Non-2G follow rules (silent)
            pass
        # propagate DataFrame changes back to final_results so writers use updated rows
        try:
            final_results = df_final.to_dict('records')
        except Exception:
            pass

except Exception:
    pass

# ------------------------------------------------------------------
# Mis-configuration detection and cleanup
# New Rule: A site must not repeat within a single traced route for a subject site.
# If a repeated site is detected, keep only the last occurrence and delete
# all sites from the first occurrence up to (but not including) the last occurrence.
# Mark the route as 'Incomplete' and set comment to 'Mis-Configured Route'.
# Record a discrepancy entry for audit in the diagnostics workbook.
def _row_path_from_record(r):
    path = []
    ne = r.get('NE')
    if pd.notna(ne):
        path.append(ne)
    i = 1
    while True:
        col = f'FE{i}'
        if col not in r or pd.isna(r.get(col)) or r.get(col) == '':
            break
        path.append(r.get(col))
        i += 1
    return path


def _compress_route(path):
    # determine first and last indices for each site
    first = {}
    last = {}
    for idx, s in enumerate(path):
        if s not in first:
            first[s] = idx
        last[s] = idx

    # build removal ranges (start_index, last_index)
    ranges = []
    for s in set(path):
        if last[s] > first[s]:
            ranges.append((first[s], last[s]))
    if not ranges:
        return path, None

    # merge overlapping ranges
    ranges.sort()
    merged = []
    cur_s, cur_e = ranges[0]
    for s, e in ranges[1:]:
        # removal spans indices [cur_s .. cur_e-1]; overlap if next start <= cur_e-1
        if s <= cur_e - 1:
            cur_e = max(cur_e, e)
        else:
            merged.append((cur_s, cur_e))
            cur_s, cur_e = s, e
    merged.append((cur_s, cur_e))

    # build set of indices to remove (from start .. end-1)
    remove_idx = set()
    for s, e in merged:
        for i in range(s, e):
            remove_idx.add(i)

    new_path = []
    removed_nodes = []
    for i, node in enumerate(path):
        if i in remove_idx:
            removed_nodes.append(node)
            continue
        new_path.append(node)
    return new_path, removed_nodes


misconf_entries = []
# Preserve traced paths from the in-memory route library for reliable
# compression logic. Ensure `route_paths` exists even if route_library
# becomes unavailable later in the script.
try:
    route_paths = {k: v.get('path') for k, v in route_library.items()}
except Exception:
    route_paths = {}
for res in final_results:
    try:
        # Prefer original traced path from route_paths when available
        key = (res.get('Site ID'), res.get('VLAN'), res.get('Service'))
        path = None
        try:
            if key in route_paths and route_paths[key]:
                path = list(route_paths[key])
        except Exception:
            path = None
        if not path:
            path = _row_path_from_record(res)
        if not path:
            continue
        new_path, removed = _compress_route(path)
        if removed:
            # Update route record: status, comment, FE hops, Drop Node
            old_comment = res.get('Comment', '') or ''
            res['Comment'] = 'Mis-Configured Route' if not old_comment else f"{old_comment} | Mis-Configured Route"
            res['Status'] = 'Incomplete'
            # update Drop Node and OFN status
            res['Drop Node'] = new_path[-1] if new_path else res.get('Drop Node')
            res['Drop Node OFN Status'] = f"{ofn_types.get(res['Drop Node'],{}).get('status','Not in OFN Sheet')} ({'LM' if ofn_types.get(res['Drop Node'],{}).get('colocation') else 'Built'})"
            # clear FE entries and set according to compressed path
            clear_fe_hops(res)
            for i, hop in enumerate(new_path[1:], start=1):
                res[f'FE{i}'] = hop
            # record discrepancy for diagnostics
            misconf_entries.append({
                'Source': 'Misconfiguration',
                'Site ID': res.get('Site ID'),
                'VLAN': res.get('VLAN'),
                'Service': res.get('Service'),
                'Removed Sequence': ','.join(removed),
                'Original Path': ' > '.join(path),
                'Optimized Path': ' > '.join(new_path),
                'Reason': 'Repeated site in route'
            })
    except Exception:
        continue

if misconf_entries:
    try:
        mismatch_df = pd.concat([mismatch_df, pd.DataFrame(misconf_entries)], ignore_index=True)
        mismatch_df = mismatch_df.drop_duplicates().sort_values(['Source', 'Site ID'])
    except Exception:
        pass

fe_cols = sorted([c for c in df_final.columns if c.startswith('FE')], key=lambda x: int(x[2:]))
cols = ['Site ID', 'Service', 'VLAN', 'Site Role', 'Tail-HUB Status', 'Site Type', 'Media Type', 'Drop Node', 'Drop Node OFN Status', 'Status', 'Comment', 'NE'] + fe_cols

df_library_state = pd.DataFrame(library_state)
df_logic = pd.DataFrame(logic_verification)

output_name = f'/home/asghar_attari1211/RnD/Output_DeepTrace_{output_date}.xlsx'
map_name = f'/home/asghar_attari1211/RnD/Site_Map_{output_date}.html'
if args.dry_run or args.sample_sites > 0 or args.sample_rows > 0 or args.filter_sites:
    output_name = f'/home/asghar_attari1211/RnD/Output_DryRun_{output_date}.xlsx'
    print_progress('ℹ️ Dry-run or sample mode: writing to separate verification file', args.verbose)

if not args.dry_run:
    def build_dependency_sheet(df_final, service, max_deps=500):
        # defensive: if df_final is empty or doesn't have Service column, return blank structure
        if 'Service' not in df_final.columns:
            return pd.DataFrame(columns=['Site Name', 'Dependent Sites', 'Dependent Site Count', 'Service Type'])
        df_svc = df_final[df_final['Service'] == service].copy()
        # FE columns (sorted lowest index to highest)
        fe_cols = sorted([c for c in df_svc.columns if str(c).upper().startswith('FE')], key=lambda x: int(str(x)[2:]))
        # If no rows for this service, return empty structured sheet
        if df_svc.empty:
            return pd.DataFrame(columns=['Site Name', 'Dependent Sites', 'Dependent Site Count', 'Service Type'])

        # Build entries following the rules: iterate farthest FE down to FE1, collect (SiteName, Dependent)
        entries = []
        for fe in reversed(fe_cols):
            sub = df_svc[df_svc[fe].notna()]
            for _, r in sub.iterrows():
                site_name = normalize_site(r[fe])
                dep_site = normalize_site(r.get('NE'))
                if site_name and dep_site:
                    entries.append((site_name, dep_site))

        # After FE1, append NE self-pairs
        for _, r in df_svc.iterrows():
            ne = normalize_site(r.get('NE'))
            if ne:
                entries.append((ne, ne))

        if not entries:
            return pd.DataFrame(columns=['Site Name', 'Dependent Sites', 'Dependent Site Count', 'Service Type'])

        # For each Site Name, collect list of dependent sites preserving order of occurrence
        grouped = defaultdict(list)
        for site, dep in entries:
            if dep not in grouped[site]:
                grouped[site].append(dep)

        # Build vertical rows: each dependent appears as its own row, but ensure first row for a site contains horizontal list
        rows = []
        for site in sorted(grouped.keys()):
            deps = grouped[site]
            # ensure site itself appears first in its deps if present in deps; otherwise insert as first
            if site in deps:
                deps = [site] + [d for d in deps if d != site]
            else:
                deps = [site] + deps
            dep_count = len(deps)
            for i, dep in enumerate(deps):
                row = {'Site Name': site, 'Dependent Sites': dep, 'Dependent Site Count': dep_count, 'Service Type': service}
                if i == 0:
                    horiz = [site] + deps[1:]
                    horiz = horiz[:max_deps]
                    for j, val in enumerate(horiz):
                        if j == 0:
                            row['NE'] = val
                        else:
                            # zero-pad dependency header numbers to 3 digits (D001, D002 ...)
                            row[f'D{j:03d}'] = val
                rows.append(row)

        df_dep = pd.DataFrame(rows)
        extra_cols = [c for c in df_dep.columns if c not in ['Site Name', 'Dependent Sites', 'Dependent Site Count', 'Service Type']]
        def col_sort_key(x):
            # NE first, then D### columns by numeric suffix, then other columns
            if x == 'NE':
                return (0, 0)
            m = re.search(r"(\d+)", x)
            if x.startswith('D') and m:
                return (1, int(m.group(1)))
            return (2, x)
        ordered = ['Site Name', 'Dependent Sites', 'Dependent Site Count', 'Service Type'] + sorted(extra_cols, key=col_sort_key)
        df_dep = df_dep.reindex(columns=ordered)
        return df_dep

    # Note: consolidated VLAN report write is postponed until after the
    # main workbook export to allow testing DeepTrace export first.

    # delete large intermediate structures to free memory
    # Keep topology and mapping structures required later (phys_graph, lr_sheet, ofn_types, fttt_sites, vsat_sites, site_coords, sl_labels)

    for var in ['vlan_db', 'vlan_presence_map', 'direct_nms', 'reverse_nms', 'evidence_by_far_end', 'ofn_df', 'fttt_df', 'vsat_df', 'sites_df', 'sl_df', 'sl_sheet']:
        try:
            if var in globals():
                del globals()[var]
        except Exception:
            pass
    try:
        gc.collect()
    except Exception:
        pass

    # Write three separate workbooks: Routes, Dependencies, Logs/Diagnostics
    routes_file = f'/home/asghar_attari1211/RnD/Routes_{output_date}.xlsx'
    deps_file = f'/home/asghar_attari1211/RnD/Dependencies_{output_date}.xlsx'
    logs_file = f'/home/asghar_attari1211/RnD/Logs_Discrepancies_{output_date}.xlsx'
    try:
        # Stream-write Final Routes using openpyxl write-only mode to reduce memory pressure
        try:
            # prepare headers
            base_cols = ['Site ID', 'Service', 'VLAN', 'Site Role', 'Tail-HUB Status', 'Site Type', 'Media Type', 'Drop Node', 'Drop Node OFN Status', 'Status', 'Comment', 'NE']
            fe_indices = set()
            for r in final_results:
                for k in r.keys():
                    if isinstance(k, str) and k.upper().startswith('FE'):
                        try:
                            fe_indices.add(int(k[2:]))
                        except Exception:
                            continue
            fe_cols = [f'FE{i}' for i in sorted(fe_indices)]

            wb = openpyxl.Workbook(write_only=True)
            ws = wb.create_sheet(title='Final Routes')
            header = base_cols + fe_cols
            ws.append(header)

            # write rows
            for r in final_results:
                row = [r.get(c, '') for c in base_cols]
                for fc in fe_cols:
                    row.append(r.get(fc, ''))
                ws.append(row)

            wb.save(routes_file)
        except MemoryError:
            print('❌ MemoryError while streaming Final Routes sheet. Consider running with --sample flags or increasing memory.')
            raise
        except Exception as e:
            print(f'⚠️ Failed to write Final Routes file: {e}')
            raise

        # Build and write Dependencies workbook (one sheet per service) using streaming
        try:
            def _write_df_to_wb_sheet(wb, sheet_name, df):
                try:
                    ws = wb.create_sheet(title=sheet_name[:31])
                except Exception:
                    ws = wb.create_sheet()
                # if df has columns, write header
                try:
                    cols = list(df.columns)
                except Exception:
                    cols = []
                if cols:
                    ws.append([str(c) for c in cols])
                    for row in df.itertuples(index=False, name=None):
                        ws.append(list(row))
                else:
                    # no columns; nothing to write but keep empty sheet
                    pass

            wb_dep = openpyxl.Workbook(write_only=True)
            # remove default sheet if present and empty
            try:
                if 'Sheet' in wb_dep.sheetnames:
                    std = wb_dep['Sheet']
                    try:
                        wb_dep.remove(std)
                    except Exception:
                        pass
            except Exception:
                pass

            for svc in ['2G', '3G', '4G', '5G']:
                try:
                    df_dep = build_dependency_sheet(pd.DataFrame(final_results), svc)
                    sheet_name = f"{svc} Dpnd"
                    _write_df_to_wb_sheet(wb_dep, sheet_name, df_dep)
                except Exception as e:
                    print(f'⚠️ Failed to build/write dependency sheet for {svc}: {e}')

            wb_dep.save(deps_file)
        except Exception as e:
            print(f'❌ Unexpected error while writing dependencies workbook: {e}')
            raise

        # Write Logs and Diagnostics workbook using streaming
        try:
            wb_logs = openpyxl.Workbook(write_only=True)
            try:
                if 'Sheet' in wb_logs.sheetnames:
                    std = wb_logs['Sheet']
                    try:
                        wb_logs.remove(std)
                    except Exception:
                        pass
            except Exception:
                pass

            try:
                _write_df_to_wb_sheet(wb_logs, 'Route Library State', df_library_state)
            except Exception as e:
                print(f'⚠️ Failed to write Route Library State: {e}')
            try:
                _write_df_to_wb_sheet(wb_logs, 'Logic Verification', df_logic)
            except Exception as e:
                print(f'⚠️ Failed to write Logic Verification: {e}')
            try:
                _write_df_to_wb_sheet(wb_logs, 'Mismatch Diagnostics', mismatch_df)
            except Exception as e:
                print(f'⚠️ Failed to write Mismatch Diagnostics: {e}')

            wb_logs.save(logs_file)
        except Exception as e:
            print(f'❌ Unexpected error while writing logs/diagnostics workbook: {e}')
            raise
    except MemoryError:
        print('❌ MemoryError while writing output files. Consider running with --sample flags or increasing memory.')
        raise
    except Exception as e:
        print(f'❌ Unexpected error while writing output files: {e}')
        raise

# Report successful save of the output workbook (similar to consolidated VLAN report message)
try:
    if os.path.exists(output_name):
        size = os.path.getsize(output_name)
        if size > 0:
            print(f'📦 Output workbook saved: {output_name} ({size//1024} KB)')
        else:
            print(f'⚠️ Output workbook created but is empty: {output_name}')
except Exception:
    pass

# Now write consolidated VLAN report (deferred) if requested
try:
    if consolidated_output and 'vlan_report' in globals():
        try:
            df_consolidated_mod = vlan_report.copy()
            # Swap labels: treat VLAN Service Name as Near End, NE NAME as Far End
            if 'VLAN Service Name' in df_consolidated_mod.columns:
                df_consolidated_mod['Near End'] = df_consolidated_mod['VLAN Service Name'].apply(lambda x: extract_sites(x)[0] if extract_sites(x) else '')
            else:
                df_consolidated_mod['Near End'] = ''
            if 'NE NAME' in df_consolidated_mod.columns:
                df_consolidated_mod['Far End'] = df_consolidated_mod['NE NAME'].apply(lambda x: extract_sites(x)[0] if extract_sites(x) else '')
            else:
                df_consolidated_mod['Far End'] = ''
            df_consolidated_mod['Mod A-B'] = df_consolidated_mod['Near End'].astype(str) + ' - ' + df_consolidated_mod['Far End'].astype(str)
            try:
                df_consolidated_mod.to_excel(consolidated_output, index=False)
                print(f'📄 Consolidated VLAN Report saved separately: {consolidated_output}')
            except Exception as e:
                print(f'⚠️ Failed to write consolidated VLAN report separately: {e}')
        except Exception:
            pass
        try:
            del vlan_report
        except Exception:
            pass
except Exception:
    pass

print(f'✅ Finished. Output workbook: {output_name} (dry-run: {args.dry_run})')
if not args.no_map:
    try:
        print(f'🗺️ Map file path: {map_name}')
    except Exception:
        pass

# Map generation disabled per request (not needed at the moment)
print_progress('ℹ️ Map generation disabled for this run', args.verbose)